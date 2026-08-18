"""
make_user_template.py — Sinh template entities.xlsx đơn giản cho user nhập.

Template gồm:
  * sheet `entities`: cột tickers/etfs/indices/exchanges/industries/entities (nhập 1 giá trị/dòng).
  * sheet `meta`: user, note.
  * sheet `huong_dan`: hướng dẫn tiếng Việt + trỏ tới data/entities/entities.xlsx để tra mã.
  * Dropdown (data-validation) cho `exchanges` (danh sách nhỏ, cố định).

Usage:
    python scripts/make_user_template.py                       # ghi users/template/entities_template.xlsx
    python scripts/make_user_template.py --seed AnPT           # tạo users/input/AnPT/entities.xlsx từ config yaml
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import yaml                                            # noqa: E402
from openpyxl import Workbook                          # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from src.core.stdio import force_utf8_stdio            # noqa: E402
from src.users.compile import (                        # noqa: E402
    DEFAULT_INPUT_ROOT, GROUP_KEYS, REPO_ROOT, USERS_CONFIG_DIR, write_user_xlsx,
)

force_utf8_stdio()

EXCHANGES = ["HOSE", "HNX", "UPCOM"]

HUONG_DAN = [
    ("Cách nhập", ""),
    ("1", "Mỗi cột trong sheet 'entities' là 1 nhóm; nhập MỖI giá trị 1 dòng."),
    ("2", "tickers/etfs/indices/exchanges: nhập MÃ (vd HPG, FPT, VNINDEX, HOSE)."),
    ("3", "industries: nhập MÃ ngành GICS (vd THEP, NGAN_HANG) — dùng code, KHÔNG dùng tên."),
    ("4", "entities: (nâng cao) entity_id nguyên bản dạng TYPE:CODE (vd TICKER:HPG)."),
    ("Tra mã", "Xem danh sách hợp lệ trong data/entities/entities.xlsx (sheet Securities/Industries/...)."),
    ("Bật/tắt", "Bật/tắt user ở users/input/manifest.yaml (vắng tên = mặc định BẬT)."),
]


def build_template(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "entities"
    ws.append(list(GROUP_KEYS))
    ws.append(["HPG", "E1VFVN30", "VNINDEX", "HOSE", "THEP", None])
    ws.freeze_panes = "A2"
    # dropdown cho exchanges (cột D)
    dv_ex = DataValidation(type="list", formula1='"%s"' % ",".join(EXCHANGES), allow_blank=True)
    ws.add_data_validation(dv_ex); dv_ex.add("D2:D200")

    ms = wb.create_sheet("meta")
    ms.append(["key", "value"])
    ms.append(["user", ""])
    ms.append(["note", ""])

    hd = wb.create_sheet("huong_dan")
    for row in HUONG_DAN:
        hd.append(list(row))
    hd.column_dimensions["B"].width = 90
    wb.save(path)
    return path


def seed_from_yaml(name: str) -> Path:
    """Tạo users/input/<name>/entities.xlsx từ config yaml hiện có (giữ dữ liệu mẫu)."""
    yaml_path = USERS_CONFIG_DIR / f"{name}.yaml"
    doc = yaml.safe_load(yaml_path.read_text(encoding="utf-8")) or {} if yaml_path.exists() else {}
    out = Path(DEFAULT_INPUT_ROOT) / name / "entities.xlsx"
    return write_user_xlsx(out, doc, {"user": name})


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", metavar="NAME", help="seed users/input/<NAME>/entities.xlsx từ config yaml")
    ap.add_argument("--out", default=str(REPO_ROOT / "users" / "template" / "entities_template.xlsx"))
    args = ap.parse_args(argv)

    if args.seed:
        p = seed_from_yaml(args.seed)
        print(f"seeded → {p}")
    p = build_template(Path(args.out))
    print(f"template → {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
