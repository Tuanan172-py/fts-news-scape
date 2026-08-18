"""
build_entities.py — Trích xuất & chuẩn hoá DANH SÁCH THỰC THỂ (entity master list)
từ dữ liệu gốc của tổ chức, phục vụ lớp L3 agent nhận diện thực thể trong tin tức.

Nguồn dữ liệu (FRA - Data):
  - trading_data/market_caps.parquet, os.xlsx, indices.parquet  -> mã (ticker) + chỉ số
  - company_data/company_name.xlsx, etf_name.xlsx               -> tên doanh nghiệp + ETF
  - industry_classification/industry_classification.xlsx        -> ngành (GICS 3 cấp)

Đầu ra (project/data/entities/):
  - entities.json        : master list (1 object / thực thể) — nguồn chân lý
  - entities.csv         : bản phẳng để tra cứu nhanh
  - taxonomy.json        : cây phân loại (type + ngành GICS + sàn + chỉ số)
  - stats.json           : thống kê build

Tiêu chí (theo yêu cầu):
  * Unique   : entity_id là khoá chính duy nhất (namespaced theo type)
  * Rõ ràng  : mỗi thực thể có canonical_name + type + aliases xác định
  * Nhất quán: mọi thực thể cùng schema
  * Có thuộc tính nhận dạng: code + aliases (surface forms để match trong text)
  * Tránh dư thừa: TICKER & tên doanh nghiệp gộp làm MỘT thực thể (name là thuộc tính)

Chạy:
  python scripts/build_entities.py
  python scripts/build_entities.py --data-root "C:/.../FRA - Data" --out data/entities
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import pandas as pd

# ----------------------------------------------------------------------------
# Cấu hình đường dẫn mặc định
# ----------------------------------------------------------------------------
DEFAULT_DATA_ROOT = Path(
    r"C:\Users\An Thanh Pham\OneDrive - fpts.com.vn\FRA - Data"
)
DEFAULT_OUT = Path(__file__).resolve().parents[1] / "data" / "entities"

SCHEMA_VERSION = "1.0.0"


# ----------------------------------------------------------------------------
# Đọc dữ liệu (parquet trên OneDrive cần đọc qua BytesIO để tránh lỗi Errno 22)
# ----------------------------------------------------------------------------
def read_parquet(path: Path) -> pd.DataFrame:
    with open(path, "rb") as f:
        return pd.read_parquet(io.BytesIO(f.read()))


def read_excel(path: Path, **kw) -> pd.DataFrame:
    return pd.read_excel(path, **kw)


# ----------------------------------------------------------------------------
# Chuẩn hoá tên & sinh alias (surface forms để nhận diện trong tin tức)
# ----------------------------------------------------------------------------
# Tiền tố pháp lý phổ biến ở đầu tên doanh nghiệp VN — cắt để lấy tên rút gọn.
_LEGAL_PREFIXES = [
    r"Tổng\s+Công\s+ty\s+Cổ\s+phần",
    r"Tổng\s+CTCP",
    r"Tổng\s+Công\s+ty",
    r"Ngân\s+hàng\s+Thương\s+mại\s+Cổ\s+phần",
    r"Ngân\s+hàng\s+TMCP",
    r"Công\s+ty\s+Cổ\s+phần",
    r"CTCP",
    r"Công\s+ty\s+TNHH\s+MTV",
    r"Công\s+ty\s+TNHH\s+Một\s+Thành\s+Viên",
    r"Công\s+ty\s+TNHH",
    r"Chứng\s+chỉ\s+Quỹ\s+ETF",
    r"Chứng\s+chỉ\s+Quỹ",
    r"Quỹ\s+Đầu\s+tư",
    r"Quỹ\s+ETF",
    r"Quỹ",
]
_PREFIX_RE = re.compile(r"^(?:%s)\b[\s:.-]*" % "|".join(_LEGAL_PREFIXES), re.IGNORECASE)
# Đuôi pháp lý: "... - CTCP", "... - Công ty Cổ phần"
_SUFFIX_RE = re.compile(r"[\s]*[-–]\s*(?:CTCP|Công\s+ty\s+Cổ\s+phần)\s*$", re.IGNORECASE)
_PAREN_RE = re.compile(r"\(([^)]+)\)")
_WS_RE = re.compile(r"\s+")


def clean_name(name: str) -> str:
    return _WS_RE.sub(" ", str(name).strip())


def short_name(name: str) -> str:
    """Lõi thương hiệu: bỏ phần ngoặc + tiền tố + đuôi pháp lý ('Tập đoàn X - CTCP' → 'X')."""
    s = _WS_RE.sub(" ", _PAREN_RE.sub("", clean_name(name))).strip()
    for _ in range(2):
        new = _PREFIX_RE.sub("", s).strip()
        if new == s:
            break
        s = new
    s = re.sub(r"^Tập\s+đoàn\b[\s:.-]*", "", s, flags=re.IGNORECASE).strip()
    s = _SUFFIX_RE.sub("", s).strip()
    s = re.sub(r"[\s\-–]+$", "", s).strip()
    return s or clean_name(name)


def make_aliases(canonical: str, code: str, brand_map: dict | None = None) -> list[str]:
    """Surface form (giữ thứ tự, unique): tên đầy đủ → tên rút gọn → thương hiệu trong
    ngoặc → alias thương hiệu bổ sung tay (brand_aliases.yaml)."""
    out: list[str] = []

    def add(x: str) -> None:
        x = clean_name(x)
        if x and x not in out and not x.isdigit() and len(x) >= 2:
            out.append(x)

    add(clean_name(canonical))
    add(short_name(canonical))
    for m in _PAREN_RE.findall(str(canonical)):   # thương hiệu trong ngoặc: (CHOLIMEX)
        add(m)
    for b in (brand_map or {}).get(code, []):     # (Vietcombank), (BIDV)...
        add(b)
    return out


# ----------------------------------------------------------------------------
# Phân loại mã: STOCK / ETF / FUND-OTHER
# ----------------------------------------------------------------------------
def classify_code(code: str, etf_codes: set[str]) -> str:
    if code in etf_codes or code.startswith(("E1", "FUE", "FUC")):
        return "ETF"
    # Mã cổ phiếu VN = 3 ký tự [A-Z0-9], bắt đầu bằng chữ (vd HPG, A32, C32, L14)
    if re.fullmatch(r"[A-Z][A-Z0-9]{2}", code):
        return "STOCK"
    # còn lại: quỹ đóng, trái phiếu, chứng quyền... -> OTHER
    return "OTHER"


# ----------------------------------------------------------------------------
# Ánh xạ tĩnh: sàn & chỉ số
# ----------------------------------------------------------------------------
EXCHANGES = [
    {"code": "HOSE", "name": "Sở Giao dịch Chứng khoán TP. Hồ Chí Minh",
     "aliases": ["HOSE", "HSX", "Sở GDCK TP.HCM", "sàn HOSE", "sàn TP.HCM"]},
    {"code": "HNX", "name": "Sở Giao dịch Chứng khoán Hà Nội",
     "aliases": ["HNX", "Sở GDCK Hà Nội", "sàn Hà Nội"]},
    {"code": "UPCOM", "name": "Thị trường giao dịch cổ phiếu công ty đại chúng chưa niêm yết",
     "aliases": ["UPCOM", "UPCoM", "sàn UPCoM"]},
]

INDEX_META = {
    "VNINDEX":  {"name": "VN-Index",       "aliases": ["VN-Index", "VNINDEX", "VN Index"], "exchange": "HOSE"},
    "VN30":     {"name": "VN30-Index",     "aliases": ["VN30", "VN30-Index", "rổ VN30"],   "exchange": "HOSE"},
    "VNXALL":   {"name": "VNX Allshare",   "aliases": ["VNXALL", "VNX Allshare", "VNX-Allshare"], "exchange": "HOSE"},
    "HNXINDEX": {"name": "HNX-Index",      "aliases": ["HNX-Index", "HNXINDEX", "HNX Index"], "exchange": "HNX"},
    "HNX30":    {"name": "HNX30-Index",    "aliases": ["HNX30", "HNX30-Index", "rổ HNX30"], "exchange": "HNX"},
    "UPINDEX":  {"name": "UPCoM-Index",    "aliases": ["UPCoM-Index", "UPINDEX", "UPCOM-Index"], "exchange": "UPCOM"},
}

def _fold_diacritics(s: str) -> str:
    """Bỏ dấu tiếng Việt: 'Bất động sản' -> 'Bat dong san' (đ/Đ -> d/D)."""
    s = s.replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def slug(s: str) -> str:
    s = _fold_diacritics(str(s))
    s = re.sub(r"[^0-9A-Za-z]+", "_", s).strip("_")
    return s.upper()


def _load_brand_aliases() -> dict:
    """config/entities/brand_aliases.yaml → {code: [alias,...]}. Thiếu file/PyYAML → {}."""
    path = Path(__file__).resolve().parents[1] / "config" / "entities" / "brand_aliases.yaml"
    try:
        import yaml
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        return {str(k).strip(): list(v or []) for k, v in data.items()}
    except Exception:  # noqa: BLE001 — brand map là tùy chọn
        return {}


# ----------------------------------------------------------------------------
# Build
# ----------------------------------------------------------------------------
def build(data_root: Path, out_dir: Path) -> dict:
    td = data_root / "trading_data"
    cd = data_root / "company_data"
    ic = data_root / "industry_classification"

    brand_map = _load_brand_aliases()

    comp = read_excel(cd / "company_name.xlsx")
    etf = read_excel(cd / "etf_name.xlsx")
    mc = read_parquet(td / "market_caps.parquet")
    osx = read_excel(td / "os.xlsx")
    idx = read_parquet(td / "indices.parquet")
    ind = read_excel(ic / "industry_classification.xlsx")

    comp_name = {str(r.Code).strip(): clean_name(r.Name) for r in comp.itertuples()}
    etf_name = {str(r.Code).strip(): clean_name(r.Name) for r in etf.itertuples()}
    etf_codes = set(etf_name)

    mc_codes = set(mc.Ticker.astype(str).str.strip())
    os_codes = set(osx.Ticker.astype(str).str.strip())
    ind_codes = set(ind.ticker.astype(str).str.strip())

    # GICS mới nhất / ticker
    ind = ind.copy()
    ind["date"] = pd.to_datetime(ind["date"])
    latest = ind.sort_values("date").groupby("ticker").tail(1)
    gics = {
        str(r.ticker).strip(): (
            clean_name(r.GICS1_name), clean_name(r.GICS2_name), clean_name(r.GICS3_name)
        )
        for r in latest.itertuples()
    }

    # OS (số lượng cổ phiếu lưu hành) mới nhất / ticker
    osx = osx.copy()
    osx["Date"] = pd.to_datetime(osx["Date"])
    os_latest = osx.sort_values("Date").groupby("Ticker").tail(1)
    os_map = {str(r.Ticker).strip(): int(r.OS) for r in os_latest.itertuples() if pd.notna(r.OS)}

    entities: list[dict] = []

    # --- 1) SECURITIES: cổ phiếu / ETF / khác (union mọi nguồn có mã) -------
    all_codes = set(comp_name) | mc_codes | os_codes | ind_codes | etf_codes
    for code in sorted(all_codes):
        kind = classify_code(code, etf_codes)
        name = etf_name.get(code) or comp_name.get(code)
        if kind == "ETF":
            etype = "ETF"
            canonical = name or code
        elif kind == "STOCK":
            etype = "TICKER"
            canonical = name or code
        else:
            etype = "SECURITY_OTHER"
            canonical = name or code

        g = gics.get(code)
        attrs: dict = {}
        if g:
            attrs["gics1"], attrs["gics2"], attrs["gics3"] = g
        if code in os_map:
            attrs["shares_outstanding"] = os_map[code]
        attrs["has_market_cap"] = code in mc_codes
        attrs["listed_universe"] = True

        ent = {
            "entity_id": f"{etype}:{code}",
            "type": etype,
            "code": code,
            "canonical_name": canonical,
            "aliases": make_aliases(canonical, code, brand_map) if (name or code in brand_map) else [],
            "attributes": attrs,
            "sources": _sources_for(code, comp_name, etf_codes, mc_codes, os_codes, gics),
        }
        entities.append(ent)

    # --- 2) INDEX: chỉ số ----------------------------------------------------
    for code in sorted(set(idx.Ticker.astype(str).str.strip())):
        meta = INDEX_META.get(code, {"name": code, "aliases": [code], "exchange": None})
        entities.append({
            "entity_id": f"INDEX:{code}",
            "type": "INDEX",
            "code": code,
            "canonical_name": meta["name"],
            "aliases": meta["aliases"],
            "attributes": {"exchange": meta["exchange"]},
            "sources": ["trading_data/indices.parquet"],
        })

    # --- 3) EXCHANGE: sàn ----------------------------------------------------
    for ex in EXCHANGES:
        entities.append({
            "entity_id": f"EXCHANGE:{ex['code']}",
            "type": "EXCHANGE",
            "code": ex["code"],
            "canonical_name": ex["name"],
            "aliases": ex["aliases"],
            "attributes": {},
            "sources": ["static/derived"],
        })

    # --- 4) INDUSTRY: ngành GICS 3 cấp --------------------------------------
    g1 = latest.groupby("GICS1_name").ticker.nunique()
    for name, n in g1.items():
        entities.append(_industry("GICS1", name, None, int(n)))
    g2 = latest.groupby(["GICS1_name", "GICS2_name"]).ticker.nunique()
    for (p1, name), n in g2.items():
        entities.append(_industry("GICS2", name, clean_name(p1), int(n)))
    g3 = latest.groupby(["GICS2_name", "GICS3_name"]).ticker.nunique()
    for (p2, name), n in g3.items():
        entities.append(_industry("GICS3", name, clean_name(p2), int(n)))

    # ----- outputs ----------------------------------------------------------
    out_dir.mkdir(parents=True, exist_ok=True)
    _write_json(out_dir / "entities.json", {
        "schema_version": SCHEMA_VERSION,
        "entity_count": len(entities),
        "entities": entities,
    })
    _write_csv(out_dir / "entities.csv", entities)
    _write_xlsx(out_dir / "entities.xlsx", entities)
    _write_taxonomy(out_dir / "taxonomy.json", entities, latest)

    stats = _stats(entities, all_codes, comp_name, gics)
    _write_json(out_dir / "stats.json", stats)
    return stats


def _sources_for(code, comp_name, etf_codes, mc_codes, os_codes, gics):
    s = []
    if code in etf_codes: s.append("company_data/etf_name.xlsx")
    if code in comp_name: s.append("company_data/company_name.xlsx")
    if code in mc_codes: s.append("trading_data/market_caps.parquet")
    if code in os_codes: s.append("trading_data/os.xlsx")
    if code in gics: s.append("industry_classification/industry_classification.xlsx")
    return s


def _industry(level: str, name: str, parent: str | None, n_tickers: int) -> dict:
    name = clean_name(name)
    return {
        "entity_id": f"IND_{level}:{slug(name)}",
        "type": f"INDUSTRY_{level}",
        "code": slug(name),
        "canonical_name": name,
        "aliases": [name],
        "attributes": {"parent": parent, "ticker_count": n_tickers},
        "sources": ["industry_classification/industry_classification.xlsx"],
    }


def _write_json(path: Path, obj) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_csv(path: Path, entities: list[dict]) -> None:
    cols = ["entity_id", "type", "code", "canonical_name", "aliases", "parent",
            "gics1", "gics2", "gics3", "sources"]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for e in entities:
            a = e["attributes"]
            w.writerow([
                e["entity_id"], e["type"], e["code"], e["canonical_name"],
                " | ".join(e["aliases"]), a.get("parent", ""),
                a.get("gics1", ""), a.get("gics2", ""), a.get("gics3", ""),
                " ; ".join(e["sources"]),
            ])


def _write_xlsx(path: Path, entities: list[dict]) -> None:
    """Workbook đa sheet theo loại — dễ phân loại/quan sát (mỗi sheet đồng nhất cột)."""

    def alias2(e):
        a = e.get("aliases") or []
        return a[1] if len(a) > 1 else (a[0] if a else "")

    sec_types = ("TICKER", "ETF", "SECURITY_OTHER")
    securities = [{
        "entity_id": e["entity_id"], "type": e["type"], "code": e["code"],
        "canonical_name": e["canonical_name"], "short_name": alias2(e),
        "gics1": e["attributes"].get("gics1", ""),
        "gics2": e["attributes"].get("gics2", ""),
        "gics3": e["attributes"].get("gics3", ""),
        "shares_outstanding": e["attributes"].get("shares_outstanding", ""),
        "has_market_cap": e["attributes"].get("has_market_cap", ""),
        "sources": " ; ".join(e["sources"]),
    } for e in entities if e["type"] in sec_types]

    industries = [{
        "entity_id": e["entity_id"], "level": e["type"].replace("INDUSTRY_", ""),
        "code": e["code"], "name": e["canonical_name"],
        "parent": e["attributes"].get("parent", ""),
        "ticker_count": e["attributes"].get("ticker_count", ""),
    } for e in entities if e["type"].startswith("INDUSTRY_")]

    indices = [{
        "entity_id": e["entity_id"], "code": e["code"], "name": e["canonical_name"],
        "aliases": " | ".join(e["aliases"]), "exchange": e["attributes"].get("exchange", ""),
    } for e in entities if e["type"] == "INDEX"]

    exchanges = [{
        "entity_id": e["entity_id"], "code": e["code"], "name": e["canonical_name"],
        "aliases": " | ".join(e["aliases"]),
    } for e in entities if e["type"] == "EXCHANGE"]

    counts = Counter(e["type"] for e in entities)
    desc = {
        "TICKER": "Cổ phiếu niêm yết (mã 3 ký tự)",
        "ETF": "Quỹ ETF / quỹ mở",
        "SECURITY_OTHER": "Quỹ đóng, trái phiếu, mã phi chuẩn",
        "INDEX": "Chỉ số thị trường",
        "EXCHANGE": "Sàn giao dịch",
        "INDUSTRY_GICS1": "Ngành GICS cấp 1", "INDUSTRY_GICS2": "Ngành GICS cấp 2",
        "INDUSTRY_GICS3": "Ngành GICS cấp 3",
    }
    index_rows = [{"type": t, "count": counts[t], "description": desc.get(t, ""),
                   "sheet": _SHEET_OF.get(t, "")} for t in counts]
    index_rows.append({"type": "TỔNG", "count": sum(counts.values()), "description": "", "sheet": ""})

    # Sheet hướng dẫn: người dùng cuối đăng ký bằng giá trị cột `code`.
    guide = pd.DataFrame([
        {"Khoá trong file đăng ký": "tickers", "Loại thực thể": "Cổ phiếu / ETF / khác",
         "Nhập giá trị ở CỘT": "code", "Sheet tra cứu": "Securities", "Ví dụ": "HPG"},
        {"Khoá trong file đăng ký": "etfs", "Loại thực thể": "Quỹ ETF",
         "Nhập giá trị ở CỘT": "code", "Sheet tra cứu": "Securities", "Ví dụ": "E1VFVN30"},
        {"Khoá trong file đăng ký": "indices", "Loại thực thể": "Chỉ số",
         "Nhập giá trị ở CỘT": "code", "Sheet tra cứu": "Indices", "Ví dụ": "VNINDEX"},
        {"Khoá trong file đăng ký": "exchanges", "Loại thực thể": "Sàn",
         "Nhập giá trị ở CỘT": "code", "Sheet tra cứu": "Exchanges", "Ví dụ": "HOSE"},
        {"Khoá trong file đăng ký": "industries", "Loại thực thể": "Ngành GICS (1/2/3)",
         "Nhập giá trị ở CỘT": "code", "Sheet tra cứu": "Industries", "Ví dụ": "THEP"},
    ], columns=["Khoá trong file đăng ký", "Loại thực thể", "Nhập giá trị ở CỘT",
                "Sheet tra cứu", "Ví dụ"])

    sheets = {
        "_Huong_dan": guide,
        "_Index": pd.DataFrame(index_rows, columns=["type", "count", "sheet", "description"]),
        "Securities": pd.DataFrame(securities),
        "Industries": pd.DataFrame(industries),
        "Indices": pd.DataFrame(indices),
        "Exchanges": pd.DataFrame(exchanges),
    }
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, df in sheets.items():
            df.to_excel(xw, sheet_name=name, index=False)
    _format_workbook(path, sheets)


_SHEET_OF = {
    "TICKER": "Securities", "ETF": "Securities", "SECURITY_OTHER": "Securities",
    "INDUSTRY_GICS1": "Industries", "INDUSTRY_GICS2": "Industries",
    "INDUSTRY_GICS3": "Industries",
    "INDEX": "Indices", "EXCHANGE": "Exchanges",
}


def _format_workbook(path: Path, sheets: dict) -> None:
    """Freeze header + AutoFilter + độ rộng cột hợp lý."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for name, df in sheets.items():
        ws = wb[name]
        ws.freeze_panes = "A2"
        ncol = max(len(df.columns), 1)
        last = ws.cell(row=1, column=ncol).column_letter
        ws.auto_filter.ref = f"A1:{last}{max(ws.max_row, 1)}"
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
            col = cell.column_letter
            width = max(12, min(48, int(df.iloc[:, c - 1].astype(str).str.len().max() if len(df) else 12) + 2))
            ws.column_dimensions[col].width = width
    wb.save(path)


def _write_taxonomy(path: Path, entities: list[dict], latest: pd.DataFrame) -> None:
    tax = {"types": dict(Counter(e["type"] for e in entities))}
    # cây GICS
    tree: dict = {}
    for (g1, g2, g3), _ in latest.groupby(["GICS1_name", "GICS2_name", "GICS3_name"]):
        tree.setdefault(clean_name(g1), {}).setdefault(clean_name(g2), set()).add(clean_name(g3))
    tax["gics_tree"] = {k: {kk: sorted(vv) for kk, vv in v.items()} for k, v in tree.items()}
    tax["exchanges"] = [e["code"] for e in EXCHANGES]
    tax["indices"] = list(INDEX_META)
    _write_json(path, tax)


def _stats(entities, all_codes, comp_name, gics) -> dict:
    by_type = Counter(e["type"] for e in entities)
    return {
        "schema_version": SCHEMA_VERSION,
        "total_entities": len(entities),
        "by_type": dict(by_type),
        "securities_total": sum(by_type[t] for t in ("TICKER", "ETF", "SECURITY_OTHER")),
        "tickers_with_name": sum(1 for c in all_codes if c in comp_name),
        "tickers_with_gics": sum(1 for c in all_codes if c in gics),
    }


def main(argv=None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser(description="Build entity master list từ dữ liệu FRA - Data")
    ap.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = ap.parse_args(argv)

    if not args.data_root.exists():
        print(f"[ERR] data-root không tồn tại: {args.data_root}", file=sys.stderr)
        return 2
    stats = build(args.data_root, args.out)
    print("== BUILD OK ==")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    print(f"\nOutput -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
