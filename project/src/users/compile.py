"""
compile.py — Biên dịch input NGƯỜI DÙNG (Excel) → config subscription (yaml).

Người dùng chỉ nhập 1 file `entities.xlsx` đơn giản trong `users/input/<name>/`. Hệ thống:
  * đọc sheet `entities` (mỗi cột = 1 nhóm: tickers/etfs/indices/exchanges/industries/entities),
  * validate qua EntityRegistry.select() (map + báo entity không tìm thấy),
  * sinh `project/config/entities/users/<name>.yaml` (định dạng máy đọc, người dùng KHÔNG đụng),
  * ghi `users/input/<name>/_unknown.txt` nếu có giá trị không map được.

Bật/tắt user đọc từ `users/input/manifest.yaml` (vắng tên = mặc định BẬT).

Không phát minh logic map — mọi ánh xạ dùng lại `EntityRegistry.select()` (entities.py).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import yaml

# src/users/compile.py → parents[2] = project/ ; repo root = project/..
PROJECT_ROOT = Path(__file__).resolve().parents[2]
REPO_ROOT = PROJECT_ROOT.parent
DEFAULT_INPUT_ROOT = REPO_ROOT / "users" / "input"
DEFAULT_OUTPUT_ROOT = REPO_ROOT / "users" / "output"
USERS_CONFIG_DIR = PROJECT_ROOT / "config" / "entities" / "users"

# Nhóm hợp lệ trong sheet `entities` — khớp đúng key của EntityRegistry.select().
GROUP_KEYS = ("tickers", "etfs", "indices", "exchanges", "industries", "entities")
# Nhóm dùng MÃ (in hoa) — chuẩn hoá để khớp entity_id (code uppercase trong entities.json).
# industries dùng CODE ngành GICS (THEP, NGAN_HANG) nên cũng viết hoa.
_UPPER_GROUPS = ("tickers", "etfs", "indices", "exchanges", "industries")


# ---- Excel I/O ----------------------------------------------------------------
def read_user_xlsx(path: str | Path) -> tuple[dict, dict]:
    """Đọc entities.xlsx → (doc, meta). doc: {group: [values]}; meta: {key: value}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    doc: dict[str, list[str]] = {}
    if "entities" in wb.sheetnames:
        rows = list(wb["entities"].iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
            col_idx = {h: i for i, h in enumerate(headers) if h in GROUP_KEYS}
            for key, idx in col_idx.items():
                vals: list[str] = []
                for r in rows[1:]:
                    if idx < len(r) and r[idx] not in (None, ""):
                        v = str(r[idx]).strip()
                        if v:
                            vals.append(v)
                if vals:
                    doc[key] = vals
    meta: dict = {}
    if "meta" in wb.sheetnames:
        for r in wb["meta"].iter_rows(values_only=True):
            if not r or not r[0]:
                continue
            key = str(r[0]).strip().lower()
            if key == "key":            # bỏ dòng tiêu đề key/value
                continue
            meta[key] = r[1] if len(r) > 1 else None
    wb.close()
    return _normalize(doc), meta


def write_user_xlsx(path: str | Path, doc: dict, meta: dict | None = None) -> Path:
    """Ghi entities.xlsx (dùng cho seed mẫu + fixture test). Cột = GROUP_KEYS."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "entities"
    ws.append(list(GROUP_KEYS))
    cols = [doc.get(k, []) for k in GROUP_KEYS]
    for i in range(max((len(c) for c in cols), default=0)):
        ws.append([c[i] if i < len(c) else None for c in cols])
    ms = wb.create_sheet("meta")
    ms.append(["key", "value"])
    for k, v in (meta or {}).items():
        ms.append([k, v])
    wb.save(path)
    return path


def _normalize(doc: dict) -> dict:
    """Chuẩn hoá: strip + upper cho nhóm dùng mã (gồm industries); giữ nguyên entities (entity_id)."""
    out: dict[str, list[str]] = {}
    for key, vals in doc.items():
        if not vals:
            continue
        if key in _UPPER_GROUPS:
            out[key] = [str(v).strip().upper() for v in vals]
        else:
            out[key] = [str(v).strip() for v in vals]
    return out


# ---- compile ------------------------------------------------------------------
_YAML_HEADER = (
    "# AUTO-GENERATED từ users/input/{name}/entities.xlsx — ĐỪNG sửa tay.\n"
    "# Chạy lại: python scripts/compile_users.py --all\n"
)


def _write_yaml(yaml_path: Path, name: str, doc: dict) -> None:
    yaml_path.parent.mkdir(parents=True, exist_ok=True)
    body = yaml.safe_dump(
        {k: doc[k] for k in GROUP_KEYS if doc.get(k)},
        allow_unicode=True, sort_keys=False, default_flow_style=False,
    )
    yaml_path.write_text(_YAML_HEADER.format(name=name) + body, encoding="utf-8")


def compile_user(name: str, xlsx_path: str | Path, registry,
                 *, users_config_dir: str | Path = USERS_CONFIG_DIR,
                 input_dir: str | Path | None = None) -> dict:
    """Compile 1 user: xlsx → yaml (+ _unknown.txt). Trả record {name, yaml_path, ids, unknown}."""
    doc, meta = read_user_xlsx(xlsx_path)
    ids, unknown = registry.select(doc)
    yaml_path = Path(users_config_dir) / f"{name}.yaml"
    _write_yaml(yaml_path, name, doc)

    if input_dir is not None:
        unk_file = Path(input_dir) / "_unknown.txt"
        if unknown:
            lines = [f"{cat}: {val} — không tìm thấy trong danh sách entity" for cat, val in unknown]
            unk_file.write_text("\n".join(lines) + "\n", encoding="utf-8")
        elif unk_file.exists():
            unk_file.unlink()  # sạch cảnh báo cũ khi user đã sửa
    return {"name": name, "yaml_path": str(yaml_path), "ids": ids,
            "unknown": unknown, "meta": meta}


def compile_all(input_root: str | Path = DEFAULT_INPUT_ROOT, registry=None,
                *, users_config_dir: str | Path = USERS_CONFIG_DIR) -> list[dict]:
    """Quét mọi users/input/<name>/entities.xlsx (bỏ folder `_*`) → compile. Trả list record.

    Sau khi ghi yaml, xoá cache registry để lần load sau nạp subscription mới.
    """
    if registry is None:
        from src.agent.entities import load_registry
        registry = load_registry()
    input_root = Path(input_root)
    results: list[dict] = []
    if input_root.exists():
        for d in sorted(p for p in input_root.iterdir() if p.is_dir()):
            if d.name.startswith("_"):
                continue
            xlsx = d / "entities.xlsx"
            if not xlsx.exists():
                continue
            results.append(compile_user(d.name, xlsx, registry,
                                        users_config_dir=users_config_dir, input_dir=d))
    try:  # invalidate lru_cache để subscription mới có hiệu lực
        from src.agent.entities import load_registry
        load_registry.cache_clear()
    except Exception:
        pass
    return results


# ---- manifest bật/tắt user ----------------------------------------------------
def load_manifest(input_root: str | Path = DEFAULT_INPUT_ROOT) -> dict[str, bool]:
    """Đọc users/input/manifest.yaml → {name: enabled}. Thiếu file/khoá = {}."""
    path = Path(input_root) / "manifest.yaml"
    if not path.exists():
        return {}
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    users = data.get("users") or {}
    return {str(k): bool(v) for k, v in users.items()}


def enabled_users(input_root: str | Path = DEFAULT_INPUT_ROOT,
                  names: list[str] | None = None) -> set[str]:
    """Tập user BẬT: manifest quyết định; vắng tên trong manifest = mặc định BẬT.

    `names` = danh sách user ứng viên (vd tên folder input). None → suy từ manifest keys.
    """
    manifest = load_manifest(input_root)
    if names is None:
        root = Path(input_root)
        names = [p.name for p in root.iterdir()
                 if p.is_dir() and not p.name.startswith("_")] if root.exists() else []
        names = sorted(set(names) | set(manifest))
    return {n for n in names if manifest.get(n, True)}
